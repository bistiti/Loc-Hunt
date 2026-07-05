#!/usr/bin/env python3
"""Crée le tableur de suivi Loc-Hunt (feuille « Logements »).

Usage :
    pip install openpyxl
    python3 scripts/creer_tracker.py [chemin_optionnel.xlsx]

Sans argument, le fichier est créé dans ~/loc-hunt-cote-azur/loc_hunt.xlsx.
Si le fichier existe déjà, il n'est PAS écrasé (pour ne pas perdre vos annonces).
"""

import pathlib
import sys

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl manquant. Installez-le : pip install openpyxl")

COLUMNS = [
    ("Titre", 45),
    ("Plateforme", 12),
    ("URL", 50),
    ("Commune", 18),
    ("Code postal", 11),
    ("Loyer CC (€)", 12),
    ("Charges comprises", 15),
    ("Surface (m²)", 11),
    ("Type", 10),
    ("Meublé", 10),
    ("Disponible le", 14),
    ("DPE", 6),
    ("Contact", 22),
    ("Notes", 40),
    ("Statut", 14),
    ("Priorité", 10),
    ("Trouvé le", 12),
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=11)


def main() -> None:
    if len(sys.argv) > 1:
        path = pathlib.Path(sys.argv[1]).expanduser()
    else:
        path = pathlib.Path.home() / "loc-hunt-cote-azur" / "loc_hunt.xlsx"

    if path.exists():
        print(f"⚠️  {path} existe déjà — non écrasé.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logements"

    for i, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"                # en-tête figé
    ws.auto_filter.ref = ws.dimensions    # filtre auto

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"✅ Créé : {path}")


if __name__ == "__main__":
    main()
