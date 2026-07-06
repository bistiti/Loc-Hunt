#!/usr/bin/env python3
"""Met à jour le tableur de suivi Loc-Hunt — SANS code Python ad hoc.

Toutes les écritures du tableur passent par ce script fixe, ce qui permet de
n'autoriser qu'une seule commande Python dans `.claude/settings.json`
(`Bash(python3 scripts/update_tracker.py *)`) plutôt que l'exécution de Python arbitraire.

Deux modes :

  1) Lister les URLs déjà présentes (pour la déduplication) :
       python3 scripts/update_tracker.py --list-urls [chemin.xlsx]

  2) Ajouter de nouvelles annonces depuis un fichier JSON :
       python3 scripts/update_tracker.py nouvelles.json [chemin.xlsx]

     `nouvelles.json` = une liste d'objets. Clés acceptées (toutes optionnelles sauf `url`) :
       titre, plateforme, url, commune, code_postal, loyer, charges_comprises,
       surface, type, meuble, disponible, dpe, contact, notes, priorite, statut
     Les doublons d'URL sont ignorés. Statut défaut = "NOUVEAU 🔴", Trouvé le = aujourd'hui.
     La ligne est colorée selon la priorité (High=vert, Medium=jaune, Low=rouge).

Sans argument de chemin, le fichier par défaut est ~/loc-hunt-cote-azur/loc_hunt.xlsx.
"""

import datetime
import json
import pathlib
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl manquant. Installez-le : pip install openpyxl")

SHEET = "Logements"
COLUMNS = [
    ("Titre", 45), ("Plateforme", 12), ("URL", 50), ("Commune", 18),
    ("Code postal", 11), ("Loyer CC (€)", 12), ("Charges comprises", 15),
    ("Surface (m²)", 11), ("Type", 12), ("Meublé", 10), ("Disponible le", 14),
    ("DPE", 6), ("Contact", 22), ("Notes", 40), ("Statut", 14),
    ("Priorité", 10), ("Trouvé le", 12),
]
COL_NAMES = [c[0] for c in COLUMNS]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Arial", size=11)
PRIORITY_FILL = {
    "high": PatternFill("solid", fgColor="E2EFDA"),
    "medium": PatternFill("solid", fgColor="FFFFC7"),
    "low": PatternFill("solid", fgColor="FCE4D6"),
}

# Alias de clés JSON -> nom de colonne exact
ALIASES = {
    "titre": "Titre", "title": "Titre",
    "plateforme": "Plateforme", "platform": "Plateforme",
    "url": "URL", "lien": "URL",
    "commune": "Commune", "ville": "Commune",
    "code_postal": "Code postal", "cp": "Code postal", "postcode": "Code postal",
    "loyer": "Loyer CC (€)", "loyer_cc": "Loyer CC (€)", "prix": "Loyer CC (€)",
    "charges_comprises": "Charges comprises", "charges": "Charges comprises",
    "surface": "Surface (m²)", "surface_m2": "Surface (m²)",
    "type": "Type",
    "meuble": "Meublé", "meublé": "Meublé", "furnished": "Meublé",
    "disponible": "Disponible le", "dispo": "Disponible le", "disponible_le": "Disponible le",
    "dpe": "DPE",
    "contact": "Contact",
    "notes": "Notes", "note": "Notes",
    "statut": "Statut", "status": "Statut",
    "priorite": "Priorité", "priorité": "Priorité", "priority": "Priorité",
    "trouve_le": "Trouvé le", "found_on": "Trouvé le",
}

PRIORITY_NORMALISE = {
    "high": "high", "haute": "high", "élevée": "high", "elevee": "high",
    "medium": "medium", "moyenne": "medium", "moyen": "medium",
    "low": "low", "basse": "low", "faible": "low",
}


def default_path() -> pathlib.Path:
    return pathlib.Path.home() / "loc-hunt-cote-azur" / "loc_hunt.xlsx"


def resolve_path(arg: str | None) -> pathlib.Path:
    return pathlib.Path(arg).expanduser() if arg else default_path()


def ensure_workbook(path: pathlib.Path):
    if path.exists():
        wb = openpyxl.load_workbook(path)
        if SHEET not in wb.sheetnames:
            ws = wb.create_sheet(SHEET)
            _write_header(ws)
        return wb, wb[SHEET]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    _write_header(ws)
    return wb, ws


def _write_header(ws) -> None:
    for i, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def existing_urls(ws) -> set:
    url_col = COL_NAMES.index("URL") + 1
    urls = set()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=url_col).value
        if val:
            urls.add(str(val).strip())
    return urls


def normalise(listing: dict) -> dict:
    """Ramène un objet JSON (clés variées) vers un dict {NomColonne: valeur}."""
    out = {}
    for key, value in listing.items():
        col = ALIASES.get(str(key).strip().lower(), None)
        if col is None and key in COL_NAMES:
            col = key
        if col:
            out[col] = value
    return out


def cmd_list_urls(path: pathlib.Path) -> None:
    if not path.exists():
        return  # aucun fichier => aucune URL
    wb = openpyxl.load_workbook(path, read_only=True)
    if SHEET not in wb.sheetnames:
        return
    for url in sorted(existing_urls(wb[SHEET])):
        print(url)


def cmd_append(json_path: pathlib.Path, xlsx_path: pathlib.Path) -> None:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list):
        sys.exit("JSON invalide : attendu une liste d'annonces.")

    wb, ws = ensure_workbook(xlsx_path)
    known = existing_urls(ws)
    today = datetime.date.today().isoformat()
    url_col = COL_NAMES.index("URL") + 1

    added = skipped = 0
    for raw in data:
        listing = normalise(raw)
        url = str(listing.get("URL", "")).strip()
        if not url:
            skipped += 1
            continue
        if url in known:
            skipped += 1
            continue
        known.add(url)

        listing.setdefault("Statut", "NOUVEAU 🔴")
        listing.setdefault("Trouvé le", today)

        row = ws.max_row + 1
        for i, name in enumerate(COL_NAMES, 1):
            ws.cell(row=row, column=i, value=listing.get(name, ""))

        # URL cliquable
        link_cell = ws.cell(row=row, column=url_col)
        link_cell.hyperlink = url
        link_cell.font = Font(color="0563C1", underline="single")

        # Couleur selon priorité
        prio = PRIORITY_NORMALISE.get(str(listing.get("Priorité", "")).strip().lower())
        if prio:
            fill = PRIORITY_FILL[prio]
            for i in range(1, len(COL_NAMES) + 1):
                ws.cell(row=row, column=i).fill = fill
        added += 1

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print(f"[OK] Ajoutees : {added} | Ignorees (doublons/URL manquante) : {skipped} | "
          f"Total lignes : {ws.max_row - 1} | Fichier : {xlsx_path}")


def main() -> None:
    args = sys.argv[1:]
    if not args:
        sys.exit(__doc__)

    if args[0] == "--list-urls":
        cmd_list_urls(resolve_path(args[1] if len(args) > 1 else None))
        return

    json_path = pathlib.Path(args[0]).expanduser()
    if not json_path.exists():
        sys.exit(f"Fichier JSON introuvable : {json_path}")
    cmd_append(json_path, resolve_path(args[1] if len(args) > 1 else None))


if __name__ == "__main__":
    main()
